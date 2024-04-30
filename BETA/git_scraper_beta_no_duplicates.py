import os
import openpyxl
from get_power import get_power

# top level repo
repo_url = 'https://github.com/fabriziotappero/ip-cores.git'

# Path to the Excel file 
excel_path = "RTL_models.xlsx"

# name of sheet
excel_sheet = "Stable"

# ID of the column containing the repository URLs 
url_column_index = 1

# ID of the column containing the top module name
top_column_index = 7

#first row with data
first_row = 2

# last row with top name
last_row = 54

# Base directory to clone the repositories into 
#base_dir = "ip-cores"
base_dir = "test-cores"
home_dir = "/home/UFAD/calewoodward/vpower"

# Ensure the base directory exists 
if not os.path.exists(base_dir):
    os.makedirs(base_dir)

# Get workbook
wb = openpyxl.load_workbook(excel_path)

# Get worksheet
ws = wb[excel_sheet]

power_vals=[]

# for each design...
for row in range(first_row, last_row): #ws.max_row
    # Get the URL from the cell in the URL column
    branch_url = ws.cell(row=row, column=url_column_index).hyperlink.target
    top_name = ws.cell(row=row, column=top_column_index).value
    power_vals.append(top_name)
    
    # Get the last part of the URL as the repo name for folder creation 
    branch_name = branch_url.split('/')[-1]
    local_branch_path = os.path.join(base_dir, branch_name)
    
    # Attempt to clone the branch
    if not os.path.exists(local_branch_path):
        try: 
            os.system('git clone -b '+branch_name+' --single-branch '+repo_url+' '+local_branch_path)
        except  Exception as e: 
            print(f"Failed to clone {branch_name}. Error: {str(e)}")
    else:
        print(f"Repository {branch_name} already exists at {local_branch_path}")

    # obtain DC_power specs for this design
    dynamic_power = get_power(top_name, home_dir+"/"+local_branch_path, home_dir)
    print(top_name+": "+str(dynamic_power)+" [W]")
    power_vals.append(dynamic_power)

print("\nCollection:")
for i in power_vals:
    print(i)
