import os
import openpyxl

# top level repo
repo_url = 'https://github.com/fabriziotappero/ip-cores.git'

# Path to the Excel file 
excel_path = "rtl_models.xlsx"

# name of sheet
excel_sheet = "Stable"

# ID of the column containing the repository URLs 
url_column_index = 1

# Base directory to clone the repositories into 
base_dir = "ip-cores"

# Ensure the base directory exists 
if not os.path.exists(base_dir):
    os.makedirs(base_dir)

# Get workbook
wb = openpyxl.load_workbook(excel_path)

# Get worksheet
ws = wb[excel_sheet]

# Iterate rows
for row in range(2, ws.max_row):
    # Get the URL from the cell in the URL column
    branch_url = ws.cell(row=row, column=url_column_index).hyperlink.target
    
    # Get the last part of the URL as the repo name for folder creation 
    branch_name = branch_url.split('/')[-1]
    local_branch_path = os.path.join(base_dir, branch_name)
    
    # Attempt to clone the branch
    if not os.path.exists(local_branch_path):
        try: 
            os.system('git clone -b '+branch_name+' --single-branch '+repo_url+' '+local_branch_path)
        except  Exception as e: 
            print(f"Failed to clone {branch_name}. Error: {str(e)}")
    else:
        print(f"Repository {branch_name} already exists at {local_branch_path}")
